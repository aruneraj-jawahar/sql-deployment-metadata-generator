from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

from src.metadata import SQLMetadata


class ExcelWriter:

    COLUMNS = [
        "File Name",
        "Deployment Group",
        "Sequence",
        "Operation",
        "Object Type",
        "Object Name",
        "SQL Statement",
    ]

    def write(
        self,
        metadata_records: list[SQLMetadata],
        output_file: str
    ) -> Path:

        if not metadata_records:
            raise ValueError(
                "No metadata records available for Excel generation."
            )

        output_path = Path(output_file)

        output_path.parent.mkdir(
            parents=True,
            exist_ok=True
        )

        records = [
            {
                "File Name": metadata.file_name,
                "Deployment Group": metadata.deployment_group,
                "Sequence": metadata.sequence,
                "Operation": metadata.operation,
                "Object Type": metadata.object_type,
                "Object Name": metadata.object_name,
                "SQL Statement": metadata.sql_statement,
            }
            for metadata in metadata_records
        ]

        dataframe = pd.DataFrame(
            records,
            columns=self.COLUMNS
        )

        dataframe = dataframe.sort_values(
            by=[
                "Deployment Group",
                "Sequence",
            ]
        )

        dataframe.to_excel(
            output_path,
            index=False,
            startrow=3,
            sheet_name="Deployment Metadata",
            engine="openpyxl"
        )

        self._format_metadata_sheet(
            output_path,
            dataframe
        )

        self._create_summary_sheet(
            output_path,
            dataframe
        )

        return output_path

    @staticmethod
    def _format_metadata_sheet(
        output_path: Path,
        dataframe: pd.DataFrame
    ) -> None:

        workbook = load_workbook(output_path)
        worksheet = workbook["Deployment Metadata"]

        # Title
        worksheet["A1"] = "SQL Deployment Metadata"
        worksheet["A1"].font = Font(
            bold=True,
            size=16
        )

        worksheet.merge_cells("A1:G1")

        # Summary information
        worksheet["A2"] = "Generated"
        worksheet["B2"] = datetime.now().strftime(
            "%Y-%m-%d %H:%M"
        )

        worksheet["D2"] = "SQL Files"
        worksheet["E2"] = len(dataframe)

        worksheet["F2"] = "Deployment Groups"
        worksheet["G2"] = dataframe[
            "Deployment Group"
        ].nunique()

        # Header formatting
        for cell in worksheet[4]:
            cell.font = Font(
                bold=True
            )
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

        # Data formatting
        for row in worksheet.iter_rows(
            min_row=5
        ):
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True
                )

        # Excel table
        if len(dataframe) > 0:

            table_reference = (
                f"A4:G{4 + len(dataframe)}"
            )

            table = Table(
                displayName="DeploymentMetadata",
                ref=table_reference
            )

            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )

            worksheet.add_table(table)

        # Freeze headers
        worksheet.freeze_panes = "A5"

        # Column widths
        column_widths = {
            "A": 28,
            "B": 24,
            "C": 12,
            "D": 14,
            "E": 22,
            "F": 30,
            "G": 60,
        }

        for column, width in column_widths.items():
            worksheet.column_dimensions[
                column
            ].width = width

        # SQL readability
        for row in range(
            5,
            worksheet.max_row + 1
        ):
            worksheet.row_dimensions[
                row
            ].height = 60

        workbook.save(output_path)

    @staticmethod
    def _create_summary_sheet(
        output_path: Path,
        dataframe: pd.DataFrame
    ) -> None:

        workbook = load_workbook(output_path)

        if "Summary" in workbook.sheetnames:
            del workbook["Summary"]

        worksheet = workbook.create_sheet(
            "Summary"
        )

        worksheet["A1"] = "Deployment Summary"
        worksheet["A1"].font = Font(
            bold=True,
            size=16
        )

        worksheet.merge_cells("A1:D1")

        headers = [
            "Deployment Group",
            "Object Count",
            "First Sequence",
            "Last Sequence",
        ]

        for column, header in enumerate(
            headers,
            start=1
        ):
            cell = worksheet.cell(
                row=3,
                column=column,
                value=header
            )

            cell.font = Font(
                bold=True
            )

            cell.alignment = Alignment(
                horizontal="center"
            )

        grouped_data = (
            dataframe
            .groupby("Deployment Group")[
                "Sequence"
            ]
            .agg(
                Object_Count="count",
                First_Sequence="min",
                Last_Sequence="max"
            )
            .reset_index()
        )

        for row_index, row in enumerate(
            grouped_data.itertuples(
                index=False
            ),
            start=4
        ):

            worksheet.cell(
                row=row_index,
                column=1,
                value=row[0]
            )

            worksheet.cell(
                row=row_index,
                column=2,
                value=row[1]
            )

            worksheet.cell(
                row=row_index,
                column=3,
                value=row[2]
            )

            worksheet.cell(
                row=row_index,
                column=4,
                value=row[3]
            )

        if len(grouped_data) > 0:

            table_reference = (
                f"A3:D{3 + len(grouped_data)}"
            )

            table = Table(
                displayName="DeploymentSummary",
                ref=table_reference
            )

            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showRowStripes=True
            )

            worksheet.add_table(table)

        worksheet.freeze_panes = "A4"

        column_widths = {
            "A": 28,
            "B": 18,
            "C": 18,
            "D": 18,
        }

        for column, width in column_widths.items():
            worksheet.column_dimensions[
                column
            ].width = width

        workbook.save(output_path)