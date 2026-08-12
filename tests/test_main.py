from pathlib import Path

from openpyxl import load_workbook

from main import main


def test_main_generates_excel(tmp_path, monkeypatch):

    input_directory = tmp_path / "sql"
    input_directory.mkdir()

    sql_file = input_directory / "customer.sql"

    sql_file.write_text(
        """-- deployment_group: CUSTOMER_DATA
-- sequence: 10

CREATE TABLE customer (
    id INT
);
""",
        encoding="utf-8",
    )

    output_file = (
        tmp_path
        / "output"
        / "deployment_metadata.xlsx"
    )

    monkeypatch.setattr(
        "sys.argv",
        [
            "main.py",
            "--input",
            str(input_directory),
            "--output",
            str(output_file),
        ],
    )

    result = main()

    assert result == 0
    assert output_file.exists()

    workbook = load_workbook(output_file)

    assert "Deployment Metadata" in workbook.sheetnames
    assert "Summary" in workbook.sheetnames


def test_main_returns_error_for_missing_directory(
    tmp_path,
    monkeypatch,
):

    input_directory = (
        tmp_path / "does_not_exist"
    )

    output_file = (
        tmp_path
        / "output"
        / "metadata.xlsx"
    )

    monkeypatch.setattr(
        "sys.argv",
        [
            "main.py",
            "--input",
            str(input_directory),
            "--output",
            str(output_file),
        ],
    )

    result = main()

    assert result == 1
    assert not output_file.exists()


def test_main_returns_error_when_input_is_file(
    tmp_path,
    monkeypatch,
):

    input_file = tmp_path / "customer.sql"

    input_file.write_text(
        "SELECT 1;",
        encoding="utf-8",
    )

    output_file = (
        tmp_path
        / "output"
        / "metadata.xlsx"
    )

    monkeypatch.setattr(
        "sys.argv",
        [
            "main.py",
            "--input",
            str(input_file),
            "--output",
            str(output_file),
        ],
    )

    result = main()

    assert result == 1
    assert not output_file.exists()


def test_main_stops_when_validation_fails(
    tmp_path,
    monkeypatch,
):

    input_directory = tmp_path / "sql"
    input_directory.mkdir()

    invalid_sql = (
        input_directory / "invalid.sql"
    )

    invalid_sql.write_text(
        """-- deployment_group: CUSTOMER_DATA

CREATE TABLE customer (
    id INT
);
""",
        encoding="utf-8",
    )

    output_file = (
        tmp_path
        / "output"
        / "metadata.xlsx"
    )

    monkeypatch.setattr(
        "sys.argv",
        [
            "main.py",
            "--input",
            str(input_directory),
            "--output",
            str(output_file),
        ],
    )

    result = main()

    assert result == 1
    assert not output_file.exists()