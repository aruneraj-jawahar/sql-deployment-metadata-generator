from pathlib import Path

import pandas as pd
import pytest

from src.excel_writer import ExcelWriter
from src.metadata import SQLMetadata


def create_metadata(
    file_name,
    deployment_group,
    sequence,
    operation,
    object_type,
    object_name,
):

    return SQLMetadata(
        file_name=file_name,
        deployment_group=deployment_group,
        sequence=sequence,
        operation=operation,
        object_type=object_type,
        object_name=object_name,
        sql_statement=f"{operation} {object_type} {object_name}",
    )


def test_excel_file_is_created(tmp_path):

    metadata = create_metadata(
        "customer_hub.sql",
        "CUSTOMER_DATA",
        10,
        "CREATE",
        "TABLE",
        "customer_hub",
    )

    output_file = tmp_path / "deployment_metadata.xlsx"

    writer = ExcelWriter()

    result = writer.write(
        [metadata],
        output_file
    )

    assert result == output_file
    assert output_file.exists()


def test_excel_contains_expected_columns(tmp_path):

    metadata = create_metadata(
        "customer_hub.sql",
        "CUSTOMER_DATA",
        10,
        "CREATE",
        "TABLE",
        "customer_hub",
    )

    output_file = tmp_path / "deployment_metadata.xlsx"

    ExcelWriter().write(
        [metadata],
        output_file
    )

    dataframe = pd.read_excel(
    output_file,
    sheet_name="Deployment Metadata",
    header=3
)

    expected_columns = [
        "File Name",
        "Deployment Group",
        "Sequence",
        "Operation",
        "Object Type",
        "Object Name",
        "SQL Statement",
    ]

    assert list(dataframe.columns) == expected_columns


def test_excel_data_is_written_correctly(tmp_path):

    metadata = create_metadata(
        "customer_hub.sql",
        "CUSTOMER_DATA",
        10,
        "CREATE",
        "TABLE",
        "customer_hub",
    )

    output_file = tmp_path / "deployment_metadata.xlsx"

    ExcelWriter().write(
        [metadata],
        output_file
    )

    dataframe = pd.read_excel(
    output_file,
    sheet_name="Deployment Metadata",
    header=3
)

    assert dataframe.iloc[0]["File Name"] == "customer_hub.sql"
    assert dataframe.iloc[0]["Deployment Group"] == "CUSTOMER_DATA"
    assert dataframe.iloc[0]["Sequence"] == 10
    assert dataframe.iloc[0]["Operation"] == "CREATE"
    assert dataframe.iloc[0]["Object Name"] == "customer_hub"


def test_records_are_sorted_by_group_and_sequence(tmp_path):

    first = create_metadata(
        "customer_load.sql",
        "CUSTOMER_DATA",
        20,
        "INSERT",
        "TABLE",
        "customer_hub",
    )

    second = create_metadata(
        "customer_hub.sql",
        "CUSTOMER_DATA",
        10,
        "CREATE",
        "TABLE",
        "customer_hub",
    )

    output_file = tmp_path / "deployment_metadata.xlsx"

    ExcelWriter().write(
        [first, second],
        output_file
    )

    dataframe = pd.read_excel(
    output_file,
    sheet_name="Deployment Metadata",
    header=3
)

    assert list(dataframe["Sequence"]) == [10, 20]


def test_empty_metadata_raises_error(tmp_path):

    output_file = tmp_path / "deployment_metadata.xlsx"

    writer = ExcelWriter()

    with pytest.raises(
        ValueError,
        match="No metadata records available"
    ):
        writer.write(
            [],
            output_file
        )

def test_excel_header_is_frozen(tmp_path):

    metadata = create_metadata(
        "customer_hub.sql",
        "CUSTOMER_DATA",
        10,
        "CREATE",
        "TABLE",
        "customer_hub",
    )

    output_file = tmp_path / "deployment_metadata.xlsx"

    ExcelWriter().write(
        [metadata],
        output_file
    )

    from openpyxl import load_workbook

    workbook = load_workbook(output_file)
    worksheet = workbook.active

    assert worksheet.freeze_panes == "A5"


def test_excel_contains_table(tmp_path):

    metadata = create_metadata(
        "customer_hub.sql",
        "CUSTOMER_DATA",
        10,
        "CREATE",
        "TABLE",
        "customer_hub",
    )

    output_file = tmp_path / "deployment_metadata.xlsx"

    ExcelWriter().write(
        [metadata],
        output_file
    )

    from openpyxl import load_workbook

    workbook = load_workbook(output_file)
    worksheet = workbook.active

    assert len(worksheet.tables) == 1
    assert "DeploymentMetadata" in worksheet.tables